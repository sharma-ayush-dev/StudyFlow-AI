import logging
import datetime
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def extract_xlsx(path: str) -> dict:
    """
    Extracts sheet names, cell values, and tabular content from an XLSX file,
    converting rows into readable text joined by ' - '.
    """
    logger.info(f"Extracting XLSX: {path}")
    text_parts = []
    
    try:
        wb = load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = [f"Sheet: {sheet_name}"]
            
            for row in sheet.iter_rows(values_only=True):
                row_vals = []
                for val in row:
                    if val is not None:
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            row_vals.append(val.strftime('%d-%m-%Y'))
                        else:
                            row_vals.append(str(val).strip())
                if row_vals:
                    # Join column values with " - "
                    sheet_text.append(" - ".join(row_vals))
            
            if len(sheet_text) > 1:
                text_parts.append("\n".join(sheet_text))
    except Exception as e:
        logger.error(f"Failed to extract XLSX: {e}")
        raise e
        
    extracted_text = "\n\n".join(text_parts)
    return {
        "text": extracted_text,
        "source_type": "xlsx",
        "ocr_used": False,
        "metadata": {}
    }
